from typing import List

import requests
import openpyxl
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from database import DataCoin, User
import pandas as pd


from bs4 import BeautifulSoup


# класс ошибок
class AuthFail(Exception):
    pass


HEADERS = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 "
    "YaBrowser/23.3.0.2246 Yowser/2.5 Safari/537.36"
}


def authorize(username, password):
    # Создаем сессию для сохранения куки
    with requests.Session() as session:
        # Авторизуемся на сайте
        resp = session.post(
            "https://ru.ucoin.net/login",
            data={"email": username, "passwd": password, "remember": 1},
            headers=HEADERS,
            allow_redirects=False,
        )
        if resp.status_code != 302:
            raise AuthFail("неверные данные авторизации")
        # print(resp.headers.get("Location"))
        user_coin_id = "".join(filter(str.isdigit, resp.headers.get("Location")))
        print(user_coin_id, "Connected")
        return user_coin_id, session


def get_graph(telegram_id):
    graph_coin_data: List[DataCoin] = DataCoin.get_for_user(telegram_id)
    print(len(graph_coin_data))

    graph_date = []
    graph_sum = []

    # last_date = datetime.strptime(graph_coin_data[0].datetime, "%Y.%m.%d")
    last_date = datetime.now().date()

    for sublist in graph_coin_data[::1]:
        while datetime.strptime(sublist.datetime, "%Y.%m.%d").date() != last_date:
            graph_date.append(last_date.strftime("%Y.%m.%d"))
            graph_sum.append(None)
            last_date -= timedelta(days=1)

        graph_date.append(sublist.datetime)
        graph_sum.append(sublist.totla_sum)
        last_date -= timedelta(days=1)

    graph_date_last_30_days = graph_date[:30]
    graph_summ_last_30_days = graph_sum[:30]
    step = len(graph_date_last_30_days) // 15

    plt.clf()
    plt.figure(figsize=(step * 5, step * 3), dpi=100)
    plt.plot(
        graph_date_last_30_days[::-1],
        graph_summ_last_30_days[::-1],
        marker="o",
        markersize=4,
    )

    new_list = list(map(lambda x: x[5:], graph_date_last_30_days))

    plt.xticks(graph_date_last_30_days[::step], new_list[::step])

    plt.title("Стоимость коллекции, руб")

    #  Прежде чем рисовать вспомогательные линии
    #  необходимо включить второстепенные деления
    plt.minorticks_on()

    #  Определяем внешний вид линий основной сетки:
    plt.grid(which="major")

    #  Определяем внешний вид линий вспомогательной
    #  сетки:
    plt.grid(
        which="minor",
        linestyle=":",
    )

    plt.savefig(f"{telegram_id}_plot.png")
    return f"{telegram_id}_plot.png"


def refresh(telegram_id):
    user = User.get(telegram_id)
    user_coin_id, session = authorize(user.email, user.password)
    file_name = download(user_coin_id, session)
    total = file_opener(file_name)
    DataCoin(user.telegram_id, total, user_coin_id).save()
    parsing(session, user, user_coin_id)


def parsing(session, user, user_coin_id):
    response = session.get(
        url=f"https://ru.ucoin.net/uid{user_coin_id}?v=home",
        headers=HEADERS,
    )
    print(response)
    soup = BeautifulSoup(response.content, "html.parser")
    results = soup.find(id="notify-popup")

    tag_messages = results.select("a:nth-child(4) div")
    tag_swap = results.select("a:nth-child(5) div")

    new_messages_count = tag_messages[1].text if len(tag_messages) == 2 else "0"
    new_swap_count = tag_swap[1].text if len(tag_swap) == 2 else "0"
    if new_messages_count.isdigit() and new_swap_count.isdigit():
        user.new_messages = int(new_messages_count)
        user.new_swap = int(new_swap_count)
        user.save()
        print("работает")
    else:
        print("Эта хуйня tag_messages tag_swap хотела быть числом")


def download(user_coin_id: str, session: requests.Session):
    # Скачиваем файл
    response = session.get(
        # URL файла, который нужно скачать
        url=f"https://ru.ucoin.net/uid{user_coin_id}?export=xls",
        headers=HEADERS,
    )

    # Имя файла, в который нужно сохранить содержимое
    file_name = f"./users_files/{user_coin_id}_.xlsx"
    with open(file_name, "wb") as f:
        f.write(response.content)

    response2 = session.get(
        # URL файла, который нужно скачать
        url=f"https://ru.ucoin.net/swap-list/?uid={user_coin_id}&export=xls",
        headers=HEADERS,
    )
    # Имя файла, в который нужно сохранить содержимое
    file_name2 = f"./users_files/{user_coin_id}_SWAP.xlsx"
    with open(file_name2, "wb") as f:
        f.write(response2.content)

    return file_name


def more_info(file_name):
    df = pd.read_excel(file_name)
    countryroad = df[df.columns[0]].unique()  # эта переменная считает кол-во стран

    return len(df), len(countryroad)


def countries(file_name):
    df = pd.read_excel("./config/RutoEng.xlsx", header=None)  # assuming no header
    mydict = df.set_index(0)[
        1
    ].to_dict()  # setting first column as index and second column as values
    df = pd.read_excel("./config/RutoCode.xlsx", header=None)  # assuming no header
    mydict1 = df.set_index(0)[1].to_dict()

    df = pd.read_excel(file_name)
    result = []
    grouped = df.groupby("Страна").size()
    for country, count in grouped.items():
        # result += f"{mydict1[country]} {str(count):<5}{country}\n            /{mydict[country]}\n"
        result.append(
            [
                mydict1[country],  # Страна
                count,  # номинал
                country,  # ГОД
                mydict[country],
            ]
        )
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    count_euro = 0
    for row in ws.iter_rows(min_row=1, max_col=7):
        if "евро" in row[1].value:
            count_euro += 1

    # result += f'🇪🇺  {count_euro}   Евросоюз\n             /Europe'
    result.append(
        [
            f"🇪🇺",
            count_euro,
            f"Евросоюз",
            f"Europe",
        ]
    )
    return result


def euro(file_name):
    df = pd.read_excel("./config/RutoCode.xlsx", header=None)  # assuming no header
    mydict1 = df.set_index(0)[1].to_dict()

    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    euros = []

    for row in ws.iter_rows(min_row=1, max_col=7):
        if "евро" in row[1].value:
            des3 = (
                f"\nМонетный двор: {row[3].value}" if row[3].value else ""
            )  # монетный двор
            des4 = f"\n{row[4].value}" if row[4].value else ""  # Наименование

            euros.append(
                [
                    f"🇪🇺 {mydict1[row[0].value]}",  # Страна
                    row[1].value,  # номинал
                    row[2].value,  # ГОД
                    f"{row[6].value} ₽",
                    des3,
                    des4,
                ]
            )

    return euros


def strana(file_name, text_in):
    # Импортируем модули для работы с Excel и ввода данных
    df = pd.read_excel("./config/RutoCode.xlsx", header=None)  # assuming no header
    mydict1 = df.set_index(0)[1].to_dict()
    # Открываем файл Excel с именем data.xlsx
    # Выбираем первый лист в файле
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Вводим текст для сравнения
    text = text_in
    print(text_in)
    string_without_first_char = text[1:]
    print(string_without_first_char)
    df = pd.read_excel("./config/EngtoRu.xlsx", header=None)  # assuming no header
    mydict = df.set_index(0)[
        1
    ].to_dict()  # setting first column as index and second column as values
    text2 = mydict[string_without_first_char]
    print(text2)

    arr = []

    # Проходимся по строкам и суммируем значения в столбце G
    for row in ws.iter_rows(min_row=1, max_col=7):
        if row[0].value == text2:
            desc3 = (
                f"\nМонетный двор: {row[3].value}" if row[3].value else ""
            )  # монетный двор
            desc4 = f"\n{row[4].value}" if row[4].value else ""  # Наименование
            arr.append(
                [
                    mydict1[row[0].value],
                    row[1].value,
                    row[2].value,
                    f"{row[6].value} ₽",
                    desc3,
                    desc4,
                ]
            )

    return arr


def func_swap(file_name):
    # Импортируем модули для работы с Excel и ввода данных
    df = pd.read_excel("./config/RutoCode.xlsx", header=None)  # assuming no header
    mydict1 = df.set_index(0)[1].to_dict()
    # Открываем файл Excel с именем data.xlsx
    # Выбираем первый лист в файле
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    arr = []

    # Проходимся по строкам и суммируем значения в столбце G
    for row in ws.iter_rows(min_row=2, max_col=11):
        desc4 = f"{row[4].value}" if row[4].value else ""  # Наименование
        desc3 = f"{row[3].value}" if row[3].value else ""  # монетный двор
        desc10 = (
            f"\nКомментарий: {row[10].value}" if row[10].value else ""
        )  # комментарий
        arr.append(
            [
                mydict1[row[0].value],  # Флаг
                row[1].value,  # Номинал
                row[2].value,  # Год
                f" {row[6].value} ₽",  # Цена
                f"\nКол-во: {row[7].value}",  # Кол-во
                desc3,  # монетный двор
                desc4,  # Наименование
                desc10,  # комментарий
            ]
        )
    return arr


def file_opener(file_name):
    # Открываем файл Excel с помощью openpyxl
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    total = 0

    # Проходимся по строкам и суммируем значения в столбце G
    for row in ws.iter_rows(min_row=2, max_col=9):
        # print(row[6].value)
        if not row[6].value:
            continue

        if row[8].value != "Метка 13":
            total += row[6].value

    return round(total, 2)
