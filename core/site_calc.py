import random
import re
from datetime import datetime, timedelta
from typing import List, Optional

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import requests
from matplotlib import ticker
from requests.exceptions import RequestException
from bs4 import BeautifulSoup

from database import DataCoin, User
from .name_transformer import transformer


# класс ошибок
class AuthFail(Exception):
    pass


# HEADERS = {
#     "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 "
#     "YaBrowser/23.3.0.2246 Yowser/2.5 Safari/537.36"
# }
HEADERS = [
    "Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; rv:109.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:109.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36 OPR/106.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36 OPR/106.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.43 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36 OPR/106.0.0.0",
]


def authorize(username, password):
    # Создаем сессию для сохранения куки
    with requests.Session() as session:
        # Авторизуемся на сайте
        resp = session.post(
            "https://ru.ucoin.net/login",
            data={"email": username, "passwd": password, "remember": 1},
            headers={"user-agent": random.choice(HEADERS)},
            allow_redirects=False,
        )
        print(datetime.now(), "| ", "Код ответа от сайта: ", resp.status_code)
        # print(datetime.now(), "| ", "С", resp.headers.get("Location"))

        if resp.status_code != 302:
            raise RequestException(f"Неверные данные авторизации. Status code: {resp.status_code}")
        # Находит в строке /uid34693?v=home цифры
        user_coin_id = "".join(filter(str.isdigit, resp.headers.get("Location")))

        print(datetime.now(), "| ", username, f"UserCoinId=[{user_coin_id}]", "Connected and authorize")
        return user_coin_id, session


def get_fig_width(data_length: int) -> int:
    if data_length > 90:
        return 20
    return 15 * (data_length // 60 or 1)


def get_date_annotation(date_value: str, data_length: int) -> str:
    date_ = datetime.strptime(date_value, "%Y.%m.%d")

    if data_length > 350:
        date_str = date_.strftime("%b %Y")
    elif data_length > 30 * 6:
        date_str = date_.strftime("%d %b")
    else:
        date_str = date_.strftime("%d.%m")

    return date_str


def get_fig_marker(data_length: int) -> str:
    if data_length < 61:
        return "o"
    return ""


def get_graph(telegram_id, limit: Optional[int] = 30):
    owne1r = User.get(telegram_id)

    len_active = len(DataCoin.get_for_user(telegram_id, limit))

    graph_coin_data: List[DataCoin] = DataCoin.get_for_user(telegram_id, limit)
    graph_date = []
    graph_sum = []
    graph_coin_count = []

    last_date = datetime.now().date()

    for sublist in graph_coin_data[::1]:
        while datetime.strptime(sublist.datetime, "%Y.%m.%d").date() != last_date:
            graph_date.append(last_date.strftime("%Y.%m.%d"))
            graph_sum.append(None)
            graph_coin_count.append(None)
            last_date -= timedelta(days=1)

        graph_date.append(sublist.datetime)
        graph_sum.append(sublist.totla_sum)
        graph_coin_count.append(sublist.totla_count)
        last_date -= timedelta(days=1)

    if limit:
        graph_date = graph_date[:limit]
        graph_sum = graph_sum[:limit]
        graph_coin_count = graph_coin_count[:limit]

    data_length = len(graph_date)

    step = data_length // 15 or 1

    fig_height = 10
    fig_width = get_fig_width(data_length)
    fig_dpi = 100

    plt.clf()
    fig, ax1 = plt.subplots(figsize=(fig_width, fig_height), dpi=fig_dpi)

    ax2 = ax1.twinx()

    ax1.plot(
        graph_date[::-1],
        graph_sum[::-1],
        marker=get_fig_marker(data_length),
        color='#0698FE',
        markersize=5,
    )
    ax2.plot(
        graph_date[::-1],
        graph_coin_count[::-1],
        marker=get_fig_marker(data_length),
        markersize=3,
        color='#30BA8F',
        linewidth=0.7
    )

    filtered_sum = [x for x in graph_sum if x is not None]  # filtered_sum равно [10, 20, 40, 50]
    maxim = max(filtered_sum)
    mimin = min(filtered_sum)
    average = sum(filtered_sum) / len(filtered_sum)
    average = round(average, 2)
    last = graph_sum[0]
    date = datetime.now
    date1 = date().strftime("%d.%m.%Y %H:%M")

    filtered_count = [x for x in graph_coin_count if x is not None]
    max_count = max(filtered_count)
    min_count = min(filtered_count)
    raznica = max_count-min_count
    y_min = min_count - 2
    y_max = min_count + raznica+2
    ax2.set_ylim(y_min, y_max)

    date_without_year = list(map(lambda value: get_date_annotation(value, data_length), graph_date))

    ax2.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    plt.xticks(graph_date[::step], date_without_year[::step])

    plt.title("Стоимость коллекции, руб")

    # Добавить текст на график
    plt.text(0, 1.07, " {}".format(owne1r.user_name), transform=plt.gca().transAxes)
    plt.text(0, 1.05, " {}".format(date1), transform=plt.gca().transAxes)

    plt.text(-0.08, 1.02, "Стоимость", color="#0698FE",  transform=plt.gca().transAxes)
    plt.text(1.02, 1.04, "Кол-во", color="#30BA8F", transform=plt.gca().transAxes)
    plt.text(1.02, 1.02, "монет", color="#30BA8F", transform=plt.gca().transAxes)

    plt.text(
        0,
        -0.1,
        "[◉_◉] Минимум = {} р.".format(mimin),
        color="red",
        transform=plt.gca().transAxes
    )
    plt.text(
        0.2,
        -0.1,
        "(◕‿◕) Максимум = {} р.".format(maxim),
        color="green",
        transform=plt.gca().transAxes,
    )
    plt.text(
        0.4,
        -0.1,
        "(─‿‿─) Средняя = {} р.".format(average),
        color="brown",
        transform=plt.gca().transAxes,
    )
    plt.text(
        0.8,
        -0.1,
        "(• ◡•) Последняя = {} р.".format(last),
        color="blue",
        transform=plt.gca().transAxes,
    )

    #  Прежде чем рисовать вспомогательные линии
    #  необходимо включить второстепенные деления
    ax1.minorticks_on()

    #  Определяем внешний вид линий основной сетки:
    ax1.grid(which="major")

    #  Определяем внешний вид линий вспомогательной
    #  сетки:
    ax1.grid(
        which="minor",
        linestyle=":",
    )

    plt.savefig(f"{telegram_id}_plot.png")
    path = f"{telegram_id}_plot.png"
    return path, len_active


def refresh(telegram_id):
    user = User.get(telegram_id)
    user_coin_id, session = authorize(user.email, user.password)
    parsing(session, user, user_coin_id)
    file_name = download(user_coin_id, session)
    total, total_count = file_opener(file_name)
    DataCoin(user.telegram_id, total, total_count).save()


def parsing(session, user, user_coin_id):
    print(datetime.now(), "| ", "Start parsing")
    try:
        response = session.get(
            url=f"https://ru.ucoin.net/uid{user_coin_id}?v=home",
            headers={"user-agent": random.choice(HEADERS)},
        )
        if response.status_code == 504:
            print(datetime.now(), "| ", f"Парсинг - ERROR: 504")

        elif response.status_code != 200:
            raise AuthFail(f"Получили ответ от сервера {response.status_code}")

    except Exception as exc:
        print(datetime.now(), "| ", f"Парсинг - ERROR: {exc}")

    else:
        soup = BeautifulSoup(response.content, "html.parser")
        mydivs = str(soup.find_all("a", {"class": "btn-s btn-gray"}))

        swap_pattern = r'<a class="btn-s btn-gray" href="/swap-mgr".*?>Обмен.*?<span class="blue-12">\(\+(\d+)\)</span></a>'
        message_pattern = r'<a class="btn-s btn-gray" href="/messages".*?>Сообщения.*?<span class="blue-12">\(\+(\d+)\)</span></a>'

        swap_matches = re.findall(swap_pattern, mydivs)
        message_matches = re.findall(message_pattern, mydivs)

        swap_count = sum(int(count) for count in swap_matches)
        message_count = sum(int(count) for count in message_matches)

        user.last_refresh = datetime.now().strftime("%d.%m.%Y %H:%M")
        user.new_messages = message_count
        user.new_swap = swap_count
        user.save()
        print(datetime.now(), "| ", f"Спарсил данные для {user.email}")


def download(user_coin_id: str, session: requests.Session):
    # Скачиваем файл
    response = session.get(
        # URL файла, который нужно скачать
        url=f"https://ru.ucoin.net/uid{user_coin_id}?export=xls",
        headers={"user-agent": random.choice(HEADERS)},
    )

    # Имя файла, в который нужно сохранить содержимое
    file_name = f"./users_files/{user_coin_id}_.xlsx"
    with open(file_name, "wb") as f:
        f.write(response.content)

    response2 = session.get(
        # URL файла, который нужно скачать
        url=f"https://ru.ucoin.net/swap-list/?uid={user_coin_id}&export=xls",
        headers={"user-agent": random.choice(HEADERS)},
    )
    # Имя файла, в который нужно сохранить содержимое
    file_name2 = f"./users_files/{user_coin_id}_SWAP.xlsx"
    with open(file_name2, "wb") as f:
        f.write(response2.content)

    return file_name


def more_info(file_name):
    # df = pd.read_excel(file_name)
    # countryroad = df[df.columns[0]].unique()  # эта переменная считает кол-во стран
    df = 0
    countryroad = 0
    sold = 0
    try:
        df = pd.read_excel(file_name)
        countryroad = df[df.columns[0]].unique()  # эта переменная считает кол-во стран
        # Получить сумму элементов в 7 столбце
        sold = df.iloc[:, 13].sum()

    except Exception:
        print(datetime.now(), "| ", f"Ошибка открытия файла")

    return len(df), len(countryroad), sold


def countries(file_name):
    df = pd.read_excel(file_name)
    result = []
    grouped = df.groupby("Страна").size()
    for country, count in grouped.items():
        # result += f"{mydict1[country]} {str(count):<5}{country}\n            /{mydict[country]}\n"
        result.append(
            [
                transformer.get_country_code(country),  # Флаг страны
                count,  # Кол-во монет
                country,  # Русское название страны
                transformer.get_country_eng_short_name(
                    country
                ),  # Короткое англ. название
            ]
        )
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    count_euro = 0
    for row in ws.iter_rows(min_row=1, max_col=7):
        if "евро" in row[1].value:
            count_euro += 1

    result.append(
        [
            f"🇪🇺",
            count_euro,
            f"Евросоюз",
            f"Europe",
        ]
    )
    return result


def euro(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    euros = []

    for row in ws.iter_rows(min_row=1, max_col=17):
        if "евро" in row[1].value:
            des2 = f"{row[2].value}г." if row[2].value else ""  # год
            des3 = (
                f"\nРазновидность: {transformer.get_coin_difference(row[3].value)}"
                if row[3].value
                else ""
            )  # монетный двор
            des4 = f"\n{row[4].value}" if row[4].value else ""  # Наименование
            des5 = f"\nМоя цена: {row[13].value} ₽" if row[13].value else ""  # Моя цена
            des6 = f"\nКомментарий: {row[16].value}" if row[16].value else ""  # Комментарий
            cena = f" {row[6].value} ₽" if row[6].value else ""  # Цена

            euros.append(
                [
                    f"🇪🇺 {transformer.get_country_code(row[0].value)}",  # Страна
                    row[1].value,  # номинал
                    des2,  # ГОД
                    cena,
                    des3,  # монетный двор
                    des4,  # Наименование
                    des5,  # покупка
                    des6,  # комментарий
                ]
            )

    return euros


def strana(file_name, text_in):
    # Открываем файл Excel с именем data.xlsx
    # Выбираем первый лист в файле
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Вводим текст для сравнения
    text = text_in
    string_without_first_char = text[1:]
    # df = pd.read_excel("./config/EngtoRu.xlsx", header=None)  # assuming no header
    # mydict = df.set_index(0)[
    #     1
    # ].to_dict()  # setting first column as index and second column as values
    text2 = transformer.get_country_rus_name(string_without_first_char)
    arr = []

    argentina = {'A': 'ОМД: Тэджон, Южная Корея. 6-угольная звезда над датой',
                 'B': 'ОМД: 6-гранный цветок над датой. Правильная надпись "PROVINCIAS"',
                 'B-er': 'ОМД: 6-гранный цветок над датой. Ошибочная надпись "PROVINGIAS"',
                 'C': 'ОМД: Париж, Франция. Ромашка с 6 лепестками над датой'}

    # Проходимся по строкам и суммируем значения в столбце G
    for row in ws.iter_rows(min_row=1, max_col=17):
        if row[0].value == text2:
            desc2 = f"{row[2].value}г." if row[2].value else ""
            if row[0].value == "Аргентина":
                desc3 = (
                    f"\nРазновидность: {argentina.get(row[3].value)}"
                    if row[3].value
                    else ""
                )  # монетный двор
            else:
                desc3 = (
                    f"\nРазновидность: {transformer.get_coin_difference(row[3].value)}"
                    if row[3].value
                    else ""
                )  # монетный двор
            desc4 = f"\n{row[4].value}" if row[4].value else ""  # Наименование
            des5 = f"\nМоя цена: {row[13].value} ₽" if row[13].value else ""  # Наименование
            des6 = f"\nКомментарий: {str(row[16].value)}" if row[16].value else ""  # Комментарий
            cena = f" {row[6].value} ₽" if row[6].value else ""  # Цена

            arr.append(
                [
                    transformer.get_country_code(row[0].value),
                    row[1].value,
                    desc2,  # ГОД
                    cena,
                    desc3,
                    desc4,
                    des5,  # покупка
                    des6,
                ]
            )

    return arr


def func_swap(file_name):
    # Открываем файл Excel с именем data.xlsx
    # Выбираем первый лист в файле
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    arr = []

    # Проходимся по строкам и суммируем значения в столбце G
    for row in ws.iter_rows(min_row=2, max_col=17):
        des2 = f"{row[2].value}г." if row[2].value else ""  # год
        desc4 = f"{row[4].value}" if row[4].value else ""  # Наименование
        desc3 = (
            f"\nРазновидность: {transformer.get_coin_difference(row[3].value)}"
            if row[3].value
            else ""
        )  # монетный двор
        desc10 = f"\nКомментарий: {row[10].value}" if row[10].value else ""  # комментарий

        arr.append(
            [
                transformer.get_country_code(row[0].value),  # Флаг
                row[1].value,  # Номинал
                des2,  # ГОД
                f" {row[6].value} ₽",  # Цена
                f"\nКол-во: {row[7].value}",  # Кол-во
                desc3,  # монетный двор
                desc4,  # Наименование
                desc10,  # комментарий
            ]
        )
    return arr


def file_opener(file_name):
    # Открываем файл Excel с помощью openpyxl
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    row_count = ws.max_row-1
    total = 0

    # Проходимся по строкам и суммируем значения в столбце G
    for row in ws.iter_rows(min_row=2, max_col=9):
        # print(row[6].value)
        if not row[6].value:
            continue

        if row[8].value != "Метка 13":
            total += row[6].value

    total_r = round(total, 2)
    return total_r, row_count


def get_top_10_coin(file_name, mode):
    df = pd.read_excel(file_name)
    arr = []
    df.fillna(value="", inplace=True)

    if mode == "old":
        df = df.sort_values(by="Год", ascending=True)
    elif mode == "novelty":
        df = df.sort_values(by="Год", ascending=False)
    elif mode == "expensive_value":
        df["Цена, RUB [uCoin]"] = pd.to_numeric(df["Цена, RUB [uCoin]"], errors="coerce")
        df = df.sort_values(by="Цена, RUB [uCoin]", ascending=False)
    elif mode == "cheap_value":
        df["Цена, RUB [uCoin]"] = pd.to_numeric(df["Цена, RUB [uCoin]"], errors="coerce")
        df = df.sort_values(by="Цена, RUB [uCoin]", ascending=True)
    elif mode == "last_append":
        df = df.sort_values(by="Добавлено", ascending=False)
    elif mode == "first_append":
        df = df.sort_values(by="Добавлено", ascending=True)

    top_10 = df.head(10)

    # Проходимся по строкам и суммируем значения в столбце G
    for row in top_10.iterrows():
        desc4 = f"\n{row[1][4]}" if row[1][4] else ""  # Наименование
        desc3 = f"{row[1][3]}" if row[1][3] else ""  # монетный двор

        desc10 = f"\nКомментарий: {row[1][16]}" if row[1][16] else ""  # комментарий
        desc5 = f" {row[1][6]} ₽" if row[1][6] else ""  # Цена
        arr.append(
            [
                transformer.get_country_code(row[1][0]),  # Флаг
                row[1][1],  # Номинал
                row[1][2],  # Год
                desc5,  # Цена
                desc3,  # монетный двор
                desc4,  # Наименование
                desc10,  # комментарий
            ]
        )
    return arr
